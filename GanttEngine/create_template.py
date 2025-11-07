"""
创建甘特图Excel模板文件
"""

import pandas as pd
from datetime import datetime, timedelta
import os

def create_gantt_template():
    """创建甘特图Excel模板"""

    # 示例数据
    today = datetime.now()

    data = {
        '任务名称': [
            '项目启动',
            '需求分析',
            '系统设计',
            '前端开发',
            '后端开发',
            '数据库设计',
            '接口开发',
            '测试',
            '部署上线',
            '项目验收'
        ],
        '开始日期': [
            (today + timedelta(days=0)).strftime('%Y-%m-%d'),
            (today + timedelta(days=2)).strftime('%Y-%m-%d'),
            (today + timedelta(days=7)).strftime('%Y-%m-%d'),
            (today + timedelta(days=14)).strftime('%Y-%m-%d'),
            (today + timedelta(days=14)).strftime('%Y-%m-%d'),
            (today + timedelta(days=10)).strftime('%Y-%m-%d'),
            (today + timedelta(days=21)).strftime('%Y-%m-%d'),
            (today + timedelta(days=35)).strftime('%Y-%m-%d'),
            (today + timedelta(days=49)).strftime('%Y-%m-%d'),
            (today + timedelta(days=56)).strftime('%Y-%m-%d')
        ],
        '结束日期': [
            (today + timedelta(days=1)).strftime('%Y-%m-%d'),
            (today + timedelta(days=6)).strftime('%Y-%m-%d'),
            (today + timedelta(days=13)).strftime('%Y-%m-%d'),
            (today + timedelta(days=34)).strftime('%Y-%m-%d'),
            (today + timedelta(days=34)).strftime('%Y-%m-%d'),
            (today + timedelta(days=13)).strftime('%Y-%m-%d'),
            (today + timedelta(days=34)).strftime('%Y-%m-%d'),
            (today + timedelta(days=48)).strftime('%Y-%m-%d'),
            (today + timedelta(days=55)).strftime('%Y-%m-%d'),
            (today + timedelta(days=60)).strftime('%Y-%m-%d')
        ],
        '负责人': [
            '张三',
            '李四',
            '王五',
            '赵六',
            '钱七',
            '孙八',
            '钱七',
            '周九',
            '吴十',
            '张三'
        ],
        '进度': [
            100,
            80,
            60,
            40,
            30,
            70,
            20,
            0,
            0,
            0
        ],
        '分类': [
            '管理',
            '规划',
            '规划',
            '开发',
            '开发',
            '开发',
            '开发',
            '测试',
            '运维',
            '管理'
        ]
    }

    # 创建DataFrame
    df = pd.DataFrame(data)

    # 确保templates目录存在
    templates_dir = os.path.join(os.path.dirname(__file__), 'templates')
    os.makedirs(templates_dir, exist_ok=True)

    # 保存为Excel文件
    output_path = os.path.join(templates_dir, 'gantt_template.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='任务列表')

        # 获取工作表
        worksheet = writer.sheets['任务列表']

        # 调整列宽
        worksheet.column_dimensions['A'].width = 20  # 任务名称
        worksheet.column_dimensions['B'].width = 15  # 开始日期
        worksheet.column_dimensions['C'].width = 15  # 结束日期
        worksheet.column_dimensions['D'].width = 12  # 负责人
        worksheet.column_dimensions['E'].width = 10  # 进度
        worksheet.column_dimensions['F'].width = 12  # 分类

        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置数据行对齐
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    print(f"模板文件已创建: {output_path}")
    return output_path

if __name__ == '__main__':
    create_gantt_template()
